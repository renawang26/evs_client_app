"""
Report Generator for EVS Navigation System.
Generates comprehensive analysis reports in Excel and PDF formats.
"""

import io
import logging
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generates Excel and PDF analysis reports for SI interpretation data."""

    # Color scheme
    COLORS = {
        'primary': '#1f77b4',
        'secondary': '#ff7f0e',
        'success': '#2ca02c',
        'warning': '#d62728',
        'info': '#9467bd',
        'light_blue': '#D6EAF8',
        'light_green': '#D5F5E3',
        'light_yellow': '#FEF9E7',
        'light_red': '#FADBD8',
        'header_bg': '#2C3E50',
        'header_fg': '#FFFFFF',
    }

    def __init__(self, file_name, lang, asr_provider):
        self.file_name = file_name
        self.lang = lang
        self.asr_provider = asr_provider
        self.words_df = None
        self.segments_df = None
        self.si_results = None
        self.metrics = {}

    def load_data(self):
        """Load all data needed for report generation."""
        from save_asr_results import get_asr_results
        from db_utils import EVSDataUtils

        self.words_df, self.segments_df = get_asr_results(
            file_name=self.file_name,
            lang=self.lang,
            asr_provider=self.asr_provider
        )

        try:
            si_df = EVSDataUtils.get_si_analysis_results(
                file_name=self.file_name,
                asr_provider=self.asr_provider,
                analysis_type='quality'
            )
            if si_df is not None and not si_df.empty:
                self.si_results = si_df.iloc[0].to_dict()
        except Exception as e:
            logger.warning(f"No SI analysis results available: {e}")
            self.si_results = None

        if self.words_df is not None and not self.words_df.empty:
            self.metrics = self._compute_metrics()

        return not (self.words_df is None or self.words_df.empty)

    def _compute_metrics(self):
        """Compute fluency and speech metrics from word-level data."""
        df = self.words_df.copy()
        metrics = {}

        # Basic counts
        total_words = len(df)
        metrics['total_words'] = total_words

        if total_words == 0:
            return metrics

        # Duration
        total_duration = df['end_time'].max() - df['start_time'].min()
        metrics['total_duration_sec'] = round(total_duration, 2)
        metrics['total_duration_min'] = round(total_duration / 60, 2)

        # Speech rate (WPM)
        if total_duration > 0:
            metrics['wpm'] = round((total_words / total_duration) * 60, 1)
        else:
            metrics['wpm'] = 0

        # Confidence stats
        if 'confidence' in df.columns:
            conf = df['confidence'].dropna()
            if len(conf) > 0:
                metrics['confidence_mean'] = round(conf.mean(), 3)
                metrics['confidence_std'] = round(conf.std(), 3)
                metrics['confidence_min'] = round(conf.min(), 3)
                metrics['confidence_max'] = round(conf.max(), 3)

        # Pause analysis (gaps between consecutive words)
        df_sorted = df.sort_values(['segment_id', 'word_seq_no'])
        gaps = df_sorted['start_time'].values[1:] - df_sorted['end_time'].values[:-1]
        pauses = gaps[gaps > 0.25]  # pauses > 250ms
        metrics['total_pauses'] = len(pauses)
        if len(pauses) > 0:
            metrics['avg_pause_duration'] = round(np.mean(pauses), 3)
            metrics['max_pause_duration'] = round(np.max(pauses), 3)
            metrics['total_pause_time'] = round(np.sum(pauses), 2)
        else:
            metrics['avg_pause_duration'] = 0
            metrics['max_pause_duration'] = 0
            metrics['total_pause_time'] = 0

        # Phonation time ratio
        speaking_time = total_duration - metrics['total_pause_time']
        metrics['phonation_ratio'] = round(speaking_time / total_duration, 3) if total_duration > 0 else 0

        # Filled pauses (filler words)
        filler_en = {'um', 'uh', 'er', 'hmm', 'like', 'you know', 'erm', 'ah'}
        filler_zh = {'啊', '嗯', '呃', '哦', '那个', '就是'}
        fillers = filler_zh if self.lang == 'zh' else filler_en

        word_col = 'edit_word' if 'edit_word' in df.columns else 'word'
        filler_mask = df[word_col].str.lower().str.strip().isin(fillers)
        metrics['filled_pauses'] = int(filler_mask.sum())
        metrics['filler_rate'] = round(metrics['filled_pauses'] / total_words * 100, 1) if total_words > 0 else 0

        # Segments
        metrics['total_segments'] = df['segment_id'].nunique() if 'segment_id' in df.columns else 0

        # Per-segment speech rate
        if 'segment_id' in df.columns and self.segments_df is not None and not self.segments_df.empty:
            seg_rates = []
            for seg_id in df['segment_id'].unique():
                seg_words = df[df['segment_id'] == seg_id]
                seg_dur = seg_words['end_time'].max() - seg_words['start_time'].min()
                if seg_dur > 0:
                    seg_rates.append({
                        'segment_id': seg_id,
                        'wpm': round((len(seg_words) / seg_dur) * 60, 1),
                        'word_count': len(seg_words),
                        'duration': round(seg_dur, 2),
                        'start_time': seg_words['start_time'].min()
                    })
            metrics['segment_rates'] = seg_rates

        return metrics

    # ─── Excel Report ────────────────────────────────────────────

    def generate_excel(self):
        """Generate Excel workbook with multiple analysis sheets."""
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            self._write_summary_sheet(writer)
            self._write_transcription_sheet(writer)
            self._write_fluency_sheet(writer)
            self._write_quality_sheet(writer)
            self._write_segments_sheet(writer)

        output.seek(0)
        return output

    def _write_summary_sheet(self, writer):
        """Write summary dashboard sheet."""
        summary_data = {
            'Metric': [],
            'Value': []
        }

        # File info
        summary_data['Metric'].append('File Name')
        summary_data['Value'].append(self.file_name)
        summary_data['Metric'].append('Language')
        summary_data['Value'].append('English' if self.lang == 'en' else 'Chinese')
        summary_data['Metric'].append('ASR Provider')
        summary_data['Value'].append(self.asr_provider)
        summary_data['Metric'].append('Report Generated')
        summary_data['Value'].append(datetime.now().strftime('%Y-%m-%d %H:%M'))
        summary_data['Metric'].append('')
        summary_data['Value'].append('')

        # Speech metrics
        summary_data['Metric'].append('--- Speech Metrics ---')
        summary_data['Value'].append('')
        summary_data['Metric'].append('Total Words')
        summary_data['Value'].append(self.metrics.get('total_words', 0))
        summary_data['Metric'].append('Total Duration (min)')
        summary_data['Value'].append(self.metrics.get('total_duration_min', 0))
        summary_data['Metric'].append('Speech Rate (WPM)')
        summary_data['Value'].append(self.metrics.get('wpm', 0))
        summary_data['Metric'].append('Total Segments')
        summary_data['Value'].append(self.metrics.get('total_segments', 0))
        summary_data['Metric'].append('')
        summary_data['Value'].append('')

        # Fluency metrics
        summary_data['Metric'].append('--- Fluency Metrics ---')
        summary_data['Value'].append('')
        summary_data['Metric'].append('Phonation Time Ratio')
        summary_data['Value'].append(self.metrics.get('phonation_ratio', 0))
        summary_data['Metric'].append('Total Pauses (>250ms)')
        summary_data['Value'].append(self.metrics.get('total_pauses', 0))
        summary_data['Metric'].append('Average Pause Duration (s)')
        summary_data['Value'].append(self.metrics.get('avg_pause_duration', 0))
        summary_data['Metric'].append('Max Pause Duration (s)')
        summary_data['Value'].append(self.metrics.get('max_pause_duration', 0))
        summary_data['Metric'].append('Filled Pauses Count')
        summary_data['Value'].append(self.metrics.get('filled_pauses', 0))
        summary_data['Metric'].append('Filler Rate (%)')
        summary_data['Value'].append(self.metrics.get('filler_rate', 0))
        summary_data['Metric'].append('')
        summary_data['Value'].append('')

        # Confidence metrics
        summary_data['Metric'].append('--- ASR Confidence ---')
        summary_data['Value'].append('')
        summary_data['Metric'].append('Mean Confidence')
        summary_data['Value'].append(self.metrics.get('confidence_mean', 'N/A'))
        summary_data['Metric'].append('Std Confidence')
        summary_data['Value'].append(self.metrics.get('confidence_std', 'N/A'))
        summary_data['Metric'].append('Min Confidence')
        summary_data['Value'].append(self.metrics.get('confidence_min', 'N/A'))
        summary_data['Metric'].append('Max Confidence')
        summary_data['Value'].append(self.metrics.get('confidence_max', 'N/A'))

        # Quality metrics from SI analysis
        if self.si_results:
            summary_data['Metric'].append('')
            summary_data['Value'].append('')
            summary_data['Metric'].append('--- SI Quality Scores ---')
            summary_data['Value'].append('')
            for key in ['overall_score', 'accuracy_score', 'fluency_score',
                        'completeness_score', 'quality_level']:
                if key in self.si_results and self.si_results[key] is not None:
                    display_name = key.replace('_', ' ').title()
                    summary_data['Metric'].append(display_name)
                    summary_data['Value'].append(self.si_results[key])

        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)

        # Style the summary sheet
        ws = writer.sheets['Summary']
        self._style_header(ws)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25

        # Style section headers
        for row in ws.iter_rows(min_row=2, max_col=2):
            val = str(row[0].value) if row[0].value else ''
            if val.startswith('---'):
                row[0].font = Font(bold=True, size=11, color='2C3E50')
                row[0].value = val.replace('---', '').strip()

    def _write_transcription_sheet(self, writer):
        """Write word-level transcription data."""
        if self.words_df is None or self.words_df.empty:
            pd.DataFrame({'Note': ['No transcription data available']}).to_excel(
                writer, sheet_name='Transcription', index=False)
            return

        cols = ['word', 'edit_word', 'start_time', 'end_time', 'confidence',
                'segment_id', 'word_seq_no', 'speaker']
        available = [c for c in cols if c in self.words_df.columns]
        df_export = self.words_df[available].copy()

        # Add duration column
        if 'start_time' in df_export.columns and 'end_time' in df_export.columns:
            df_export['duration'] = (df_export['end_time'] - df_export['start_time']).round(3)

        df_export.to_excel(writer, sheet_name='Transcription', index=False)

        ws = writer.sheets['Transcription']
        self._style_header(ws)
        self._auto_width(ws)

    def _write_fluency_sheet(self, writer):
        """Write fluency profile data."""
        rows = []

        # Overall fluency metrics
        rows.append({
            'Category': 'Speech Rate',
            'Metric': 'Words Per Minute (WPM)',
            'Value': self.metrics.get('wpm', 0),
            'Assessment': self._assess_wpm(self.metrics.get('wpm', 0))
        })
        rows.append({
            'Category': 'Speech Rate',
            'Metric': 'Total Duration (minutes)',
            'Value': self.metrics.get('total_duration_min', 0),
            'Assessment': ''
        })
        rows.append({
            'Category': 'Fluency',
            'Metric': 'Phonation Time Ratio',
            'Value': self.metrics.get('phonation_ratio', 0),
            'Assessment': self._assess_ptr(self.metrics.get('phonation_ratio', 0))
        })
        rows.append({
            'Category': 'Pauses',
            'Metric': 'Total Pauses (>250ms)',
            'Value': self.metrics.get('total_pauses', 0),
            'Assessment': ''
        })
        rows.append({
            'Category': 'Pauses',
            'Metric': 'Average Pause Duration (s)',
            'Value': self.metrics.get('avg_pause_duration', 0),
            'Assessment': ''
        })
        rows.append({
            'Category': 'Pauses',
            'Metric': 'Max Pause Duration (s)',
            'Value': self.metrics.get('max_pause_duration', 0),
            'Assessment': ''
        })
        rows.append({
            'Category': 'Pauses',
            'Metric': 'Total Pause Time (s)',
            'Value': self.metrics.get('total_pause_time', 0),
            'Assessment': ''
        })
        rows.append({
            'Category': 'Fillers',
            'Metric': 'Filled Pause Count',
            'Value': self.metrics.get('filled_pauses', 0),
            'Assessment': ''
        })
        rows.append({
            'Category': 'Fillers',
            'Metric': 'Filler Rate (%)',
            'Value': self.metrics.get('filler_rate', 0),
            'Assessment': self._assess_filler_rate(self.metrics.get('filler_rate', 0))
        })

        df_fluency = pd.DataFrame(rows)
        df_fluency.to_excel(writer, sheet_name='Fluency Profile', index=False)

        ws = writer.sheets['Fluency Profile']
        self._style_header(ws)
        self._auto_width(ws)

        # Per-segment speech rate table
        seg_rates = self.metrics.get('segment_rates', [])
        if seg_rates:
            df_seg = pd.DataFrame(seg_rates)
            start_row = len(rows) + 4
            ws.cell(row=start_row, column=1, value='Per-Segment Speech Rate').font = Font(bold=True, size=12)
            df_seg.to_excel(writer, sheet_name='Fluency Profile', index=False, startrow=start_row)

    def _write_quality_sheet(self, writer):
        """Write SI quality metrics sheet."""
        if not self.si_results:
            pd.DataFrame({'Note': ['No SI quality analysis available. Run SI Analysis first.']}).to_excel(
                writer, sheet_name='Quality Metrics', index=False)
            return

        rows = []
        quality_fields = [
            ('overall_score', 'Overall Score', '/100'),
            ('accuracy_score', 'Accuracy Score', '/100'),
            ('fluency_score', 'Fluency Score', '/100'),
            ('completeness_score', 'Completeness Score', '/100'),
            ('quality_level', 'Quality Level', ''),
            ('en_wpm', 'English WPM', 'words/min'),
            ('zh_wpm', 'Chinese WPM', 'chars/min'),
            ('speed_ratio', 'Speed Ratio', ''),
            ('pace_assessment', 'Pace Assessment', ''),
            ('balance_assessment', 'Balance Assessment', ''),
            ('coverage_rate', 'Coverage Rate', '%'),
            ('total_segments', 'Total Segments', ''),
            ('bilingual_segments', 'Bilingual Segments', ''),
            ('confidence_mean', 'Confidence Mean', ''),
            ('confidence_std', 'Confidence Std', ''),
            ('total_errors', 'Total Errors', ''),
            ('error_density', 'Error Density', ''),
            ('average_delay', 'Average Delay', 's'),
            ('max_delay', 'Max Delay', 's'),
            ('sync_quality', 'Sync Quality', ''),
            ('adaptation_score', 'Cultural Adaptation Score', '/100'),
        ]

        for field, display, unit in quality_fields:
            val = self.si_results.get(field)
            if val is not None:
                rows.append({
                    'Metric': display,
                    'Value': val,
                    'Unit': unit
                })

        df_quality = pd.DataFrame(rows)
        df_quality.to_excel(writer, sheet_name='Quality Metrics', index=False)

        ws = writer.sheets['Quality Metrics']
        self._style_header(ws)
        self._auto_width(ws)

    def _write_segments_sheet(self, writer):
        """Write segment-level data."""
        if self.segments_df is None or self.segments_df.empty:
            pd.DataFrame({'Note': ['No segment data available']}).to_excel(
                writer, sheet_name='Segments', index=False)
            return

        cols = ['segment_id', 'start_time', 'end_time', 'text', 'edit_text', 'speaker']
        available = [c for c in cols if c in self.segments_df.columns]
        df_export = self.segments_df[available].copy()

        if 'start_time' in df_export.columns and 'end_time' in df_export.columns:
            df_export['duration'] = (df_export['end_time'] - df_export['start_time']).round(2)

        df_export.to_excel(writer, sheet_name='Segments', index=False)

        ws = writer.sheets['Segments']
        self._style_header(ws)
        self._auto_width(ws)

    # ─── PDF Report ──────────────────────────────────────────────

    def generate_pdf(self):
        """Generate PDF report with charts and tables."""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output, pagesize=A4,
            topMargin=2 * cm, bottomMargin=2 * cm,
            leftMargin=2 * cm, rightMargin=2 * cm
        )

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            name='ReportTitle', parent=styles['Title'],
            fontSize=20, spaceAfter=20, textColor=colors.HexColor('#2C3E50')
        ))
        styles.add(ParagraphStyle(
            name='SectionHeader', parent=styles['Heading2'],
            fontSize=14, spaceBefore=16, spaceAfter=8,
            textColor=colors.HexColor('#2C3E50'),
            borderWidth=1, borderColor=colors.HexColor('#2C3E50'),
            borderPadding=4
        ))
        styles.add(ParagraphStyle(
            name='MetricLabel', parent=styles['Normal'],
            fontSize=10, textColor=colors.HexColor('#7F8C8D')
        ))
        styles.add(ParagraphStyle(
            name='MetricValue', parent=styles['Normal'],
            fontSize=16, textColor=colors.HexColor('#2C3E50'),
            alignment=TA_CENTER
        ))

        elements = []

        # Title
        lang_display = 'English' if self.lang == 'en' else 'Chinese'
        elements.append(Paragraph('EVS Analysis Report', styles['ReportTitle']))
        elements.append(Paragraph(
            f'<b>File:</b> {self.file_name}<br/>'
            f'<b>Language:</b> {lang_display}<br/>'
            f'<b>ASR Provider:</b> {self.asr_provider}<br/>'
            f'<b>Generated:</b> {datetime.now().strftime("%Y-%m-%d %H:%M")}',
            styles['Normal']
        ))
        elements.append(Spacer(1, 10 * mm))

        # Summary metrics table
        elements.append(Paragraph('Summary', styles['SectionHeader']))
        summary_table = self._build_summary_table(colors)
        if summary_table:
            elements.append(summary_table)
        elements.append(Spacer(1, 8 * mm))

        # Charts
        charts = self._generate_charts()
        for title, img_bytes in charts:
            if img_bytes:
                elements.append(Paragraph(title, styles['SectionHeader']))
                img = Image(io.BytesIO(img_bytes), width=16 * cm, height=9 * cm)
                elements.append(img)
                elements.append(Spacer(1, 6 * mm))

        # Quality metrics (if available)
        if self.si_results:
            elements.append(PageBreak())
            elements.append(Paragraph('SI Quality Assessment', styles['SectionHeader']))
            quality_table = self._build_quality_table(colors)
            if quality_table:
                elements.append(quality_table)

        # Fluency profile table
        elements.append(Spacer(1, 8 * mm))
        elements.append(Paragraph('Fluency Profile', styles['SectionHeader']))
        fluency_table = self._build_fluency_table(colors)
        if fluency_table:
            elements.append(fluency_table)

        doc.build(elements)
        output.seek(0)
        return output

    def _build_summary_table(self, colors):
        """Build summary metrics table for PDF."""
        data = [
            ['Total Words', 'Duration (min)', 'Speech Rate (WPM)', 'Segments'],
            [
                str(self.metrics.get('total_words', 0)),
                str(self.metrics.get('total_duration_min', 0)),
                str(self.metrics.get('wpm', 0)),
                str(self.metrics.get('total_segments', 0))
            ],
            ['Pauses', 'Avg Pause (s)', 'Filler Count', 'Filler Rate (%)'],
            [
                str(self.metrics.get('total_pauses', 0)),
                str(self.metrics.get('avg_pause_duration', 0)),
                str(self.metrics.get('filled_pauses', 0)),
                str(self.metrics.get('filler_rate', 0))
            ]
        ]

        table = Table(data, colWidths=[100, 100, 110, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 2), (-1, 2), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTSIZE', (0, 1), (-1, 1), 14),
            ('FONTSIZE', (0, 3), (-1, 3), 14),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        return table

    def _build_quality_table(self, colors):
        """Build SI quality metrics table for PDF."""
        if not self.si_results:
            return None

        data = [['Metric', 'Value']]
        fields = [
            ('overall_score', 'Overall Score'),
            ('accuracy_score', 'Accuracy'),
            ('fluency_score', 'Fluency'),
            ('completeness_score', 'Completeness'),
            ('quality_level', 'Quality Level'),
            ('coverage_rate', 'Coverage Rate'),
        ]
        for field, display in fields:
            val = self.si_results.get(field)
            if val is not None:
                data.append([display, str(val)])

        if len(data) <= 1:
            return None

        table = Table(data, colWidths=[200, 200])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ]))
        return table

    def _build_fluency_table(self, colors):
        """Build fluency profile table for PDF."""
        data = [['Metric', 'Value', 'Assessment']]
        data.append(['Speech Rate (WPM)', str(self.metrics.get('wpm', 0)),
                      self._assess_wpm(self.metrics.get('wpm', 0))])
        data.append(['Phonation Ratio', str(self.metrics.get('phonation_ratio', 0)),
                      self._assess_ptr(self.metrics.get('phonation_ratio', 0))])
        data.append(['Total Pauses', str(self.metrics.get('total_pauses', 0)), ''])
        data.append(['Avg Pause (s)', str(self.metrics.get('avg_pause_duration', 0)), ''])
        data.append(['Filler Rate (%)', str(self.metrics.get('filler_rate', 0)),
                      self._assess_filler_rate(self.metrics.get('filler_rate', 0))])

        table = Table(data, colWidths=[150, 100, 150])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ]))
        return table

    # ─── Charts ──────────────────────────────────────────────────

    def _generate_charts(self):
        """Generate all charts as PNG bytes for PDF embedding."""
        charts = []

        charts.append(('Speech Rate Over Time', self._chart_speech_rate()))
        charts.append(('Confidence Distribution', self._chart_confidence()))
        charts.append(('Fluency Breakdown', self._chart_fluency_breakdown()))
        charts.append(('Pause Distribution', self._chart_pause_distribution()))

        if self.si_results:
            charts.append(('Quality Radar', self._chart_quality_radar()))

        return charts

    def _chart_speech_rate(self):
        """Line chart: speech rate per segment over time."""
        seg_rates = self.metrics.get('segment_rates', [])
        if not seg_rates:
            return None

        df = pd.DataFrame(seg_rates)
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=df['start_time'] / 60,
            y=df['wpm'],
            mode='lines+markers',
            name='WPM',
            line=dict(color=self.COLORS['primary'], width=2),
            marker=dict(size=6)
        ))

        # Add average line
        avg_wpm = self.metrics.get('wpm', 0)
        fig.add_hline(y=avg_wpm, line_dash="dash",
                       line_color=self.COLORS['secondary'],
                       annotation_text=f"Avg: {avg_wpm}")

        fig.update_layout(
            title='Speech Rate Over Time',
            xaxis_title='Time (minutes)',
            yaxis_title='Words Per Minute',
            template='plotly_white',
            height=400, width=700,
            margin=dict(l=60, r=30, t=50, b=50)
        )

        try:
            return fig.to_image(format='png', scale=2)
        except Exception as e:
            logger.warning(f"Failed to render speech rate chart: {e}")
            return None

    def _chart_confidence(self):
        """Histogram: word-level confidence distribution."""
        if self.words_df is None or 'confidence' not in self.words_df.columns:
            return None

        conf = self.words_df['confidence'].dropna()
        if len(conf) == 0:
            return None

        fig = go.Figure()
        fig.add_trace(go.Histogram(
            x=conf, nbinsx=20,
            marker_color=self.COLORS['primary'],
            opacity=0.8
        ))
        fig.add_vline(x=conf.mean(), line_dash="dash",
                       line_color=self.COLORS['warning'],
                       annotation_text=f"Mean: {conf.mean():.3f}")

        fig.update_layout(
            title='ASR Confidence Distribution',
            xaxis_title='Confidence Score',
            yaxis_title='Word Count',
            template='plotly_white',
            height=400, width=700,
            margin=dict(l=60, r=30, t=50, b=50)
        )

        try:
            return fig.to_image(format='png', scale=2)
        except Exception as e:
            logger.warning(f"Failed to render confidence chart: {e}")
            return None

    def _chart_fluency_breakdown(self):
        """Donut chart: proportion of fluent speech vs fillers vs pauses."""
        total_dur = self.metrics.get('total_duration_sec', 0)
        if total_dur <= 0:
            return None

        pause_time = self.metrics.get('total_pause_time', 0)
        speaking_time = total_dur - pause_time

        # Estimate filler time (avg word duration * filler count)
        if self.words_df is not None and not self.words_df.empty:
            avg_word_dur = (self.words_df['end_time'] - self.words_df['start_time']).mean()
            filler_time = self.metrics.get('filled_pauses', 0) * avg_word_dur
        else:
            filler_time = 0

        fluent_time = max(0, speaking_time - filler_time)

        labels = ['Fluent Speech', 'Filled Pauses', 'Silent Pauses']
        values = [round(fluent_time, 1), round(filler_time, 1), round(pause_time, 1)]
        chart_colors = [self.COLORS['success'], self.COLORS['secondary'], self.COLORS['warning']]

        fig = go.Figure(data=[go.Pie(
            labels=labels, values=values,
            hole=0.4,
            marker=dict(colors=chart_colors),
            textinfo='label+percent',
            textfont_size=12
        )])

        fig.update_layout(
            title='Time Distribution',
            template='plotly_white',
            height=400, width=700,
            margin=dict(l=30, r=30, t=50, b=30)
        )

        try:
            return fig.to_image(format='png', scale=2)
        except Exception as e:
            logger.warning(f"Failed to render fluency chart: {e}")
            return None

    def _chart_pause_distribution(self):
        """Histogram: pause duration distribution."""
        if self.words_df is None or self.words_df.empty:
            return None

        df_sorted = self.words_df.sort_values(['segment_id', 'word_seq_no'])
        gaps = df_sorted['start_time'].values[1:] - df_sorted['end_time'].values[:-1]
        pauses = gaps[gaps > 0.25]

        if len(pauses) == 0:
            return None

        fig = go.Figure()
        fig.add_trace(go.Histogram(
            x=pauses, nbinsx=15,
            marker_color=self.COLORS['info'],
            opacity=0.8
        ))
        fig.add_vline(x=np.mean(pauses), line_dash="dash",
                       line_color=self.COLORS['warning'],
                       annotation_text=f"Mean: {np.mean(pauses):.2f}s")

        fig.update_layout(
            title='Pause Duration Distribution',
            xaxis_title='Pause Duration (seconds)',
            yaxis_title='Count',
            template='plotly_white',
            height=400, width=700,
            margin=dict(l=60, r=30, t=50, b=50)
        )

        try:
            return fig.to_image(format='png', scale=2)
        except Exception as e:
            logger.warning(f"Failed to render pause chart: {e}")
            return None

    def _chart_quality_radar(self):
        """Radar chart: SI quality dimensions."""
        if not self.si_results:
            return None

        categories = ['Accuracy', 'Fluency', 'Completeness', 'Coverage', 'Adaptation']
        values = [
            self.si_results.get('accuracy_score', 0) or 0,
            self.si_results.get('fluency_score', 0) or 0,
            self.si_results.get('completeness_score', 0) or 0,
            (self.si_results.get('coverage_rate', 0) or 0) * 100 if isinstance(self.si_results.get('coverage_rate'), float) and self.si_results.get('coverage_rate', 0) < 1 else (self.si_results.get('coverage_rate', 0) or 0),
            self.si_results.get('adaptation_score', 0) or 0,
        ]

        # Filter out all-zero
        if all(v == 0 for v in values):
            return None

        fig = go.Figure()
        fig.add_trace(go.Scatterpolar(
            r=values + [values[0]],  # close the polygon
            theta=categories + [categories[0]],
            fill='toself',
            fillcolor='rgba(31, 119, 180, 0.3)',
            line=dict(color=self.COLORS['primary'], width=2),
            name='Quality'
        ))

        fig.update_layout(
            polar=dict(
                radialaxis=dict(visible=True, range=[0, 100]),
            ),
            title='SI Quality Assessment',
            template='plotly_white',
            height=450, width=700,
            margin=dict(l=60, r=60, t=60, b=40)
        )

        try:
            return fig.to_image(format='png', scale=2)
        except Exception as e:
            logger.warning(f"Failed to render quality radar chart: {e}")
            return None

    # ─── Combined Report ─────────────────────────────────────────

    def generate_full_report(self):
        """Generate both Excel and PDF reports.

        Returns:
            tuple: (excel_bytes: BytesIO, pdf_bytes: BytesIO)
        """
        excel = self.generate_excel()
        pdf = self.generate_pdf()
        return excel, pdf

    # ─── Helpers ─────────────────────────────────────────────────

    def _style_header(self, ws):
        """Apply header styling to worksheet."""
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.freeze_panes = 'A2'

    def _auto_width(self, ws):
        """Auto-adjust column widths."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    @staticmethod
    def _assess_wpm(wpm):
        if wpm == 0:
            return ''
        if wpm < 100:
            return 'Slow'
        elif wpm < 140:
            return 'Normal'
        elif wpm < 170:
            return 'Moderate-fast'
        else:
            return 'Fast'

    @staticmethod
    def _assess_ptr(ptr):
        if ptr == 0:
            return ''
        if ptr < 0.5:
            return 'Low (many pauses)'
        elif ptr < 0.7:
            return 'Normal'
        else:
            return 'High (fluent)'

    @staticmethod
    def _assess_filler_rate(rate):
        if rate == 0:
            return 'None'
        if rate < 3:
            return 'Low'
        elif rate < 6:
            return 'Moderate'
        else:
            return 'High'
